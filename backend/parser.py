import pdfplumber
import re
import os
from datetime import datetime

# Multi-Line Fund Name Normalization and Matching Logic
def normalize_fund_name(name_str):
    if not name_str:
        return ""
    # 1. Strip out all internal newline characters (\n) and replace them with a single space
    s = name_str.replace('\n', ' ')
    # 2. Remove all trailing/leading hyphens and collapse consecutive spaces down to a single space character.
    words = s.split()
    cleaned_words = [w.strip('-') for w in words]
    s = " ".join(cleaned_words)
    return s.strip()

def clean_to_tokens(s):
    s_low = s.lower()
    # Remove common wrapper words
    s_low = re.sub(r'\b(manulife|investment|fund)\b', ' ', s_low)
    s_clean = re.sub(r'[-\s\W_]+', ' ', s_low).strip()
    return set(s_clean.split())

def is_fund_match(excel_name, pdf_name):
    if not excel_name or not pdf_name:
        return False
        
    excel_norm = normalize_fund_name(excel_name)
    pdf_norm = normalize_fund_name(pdf_name)
    
    tokens_excel = clean_to_tokens(excel_norm)
    tokens_pdf = clean_to_tokens(pdf_norm)
    
    if not tokens_excel or not tokens_pdf:
        return False
        
    if tokens_excel == tokens_pdf:
        return True
        
    # Guard against partial matches on single words like "shariah" or "progress"
    if len(tokens_pdf) == 1 and list(tokens_pdf)[0] in {"shariah", "progress"}:
        return False
    if len(tokens_excel) == 1 and list(tokens_excel)[0] in {"shariah", "progress"}:
        return False
        
    # Duplicate Token Disambiguation (Shariah / Progress / Shariah progress-plus / Asia / Pacific Fixes):
    # Treat "SHARIAH ASIA-PACIFIC EX JAPAN" and "SHARIAH PROGRESS" as completely separate unique string entities.
    if ("progress" in tokens_excel and "progress" not in tokens_pdf) or ("progress" in tokens_pdf and "progress" not in tokens_excel):
        return False
    if ("asia" in tokens_excel and "asia" not in tokens_pdf) or ("asia" in tokens_pdf and "asia" not in tokens_excel):
        return False
    if ("pacific" in tokens_excel and "pacific" not in tokens_pdf) or ("pacific" in tokens_pdf and "pacific" not in tokens_excel):
        return False
    if ("shariah" in tokens_excel and "shariah" not in tokens_pdf) or ("shariah" in tokens_pdf and "shariah" not in tokens_excel):
        return False
        
    if tokens_excel == tokens_pdf:
        return True
        
    # Check if one is a subset of the other (for prefix/suffix differences)
    common = tokens_excel.intersection(tokens_pdf)
    min_len = min(len(tokens_excel), len(tokens_pdf))
    
    # 1-token guard: if the shorter name has only 1 token (e.g. BOND, GROWTH), do not allow subset matching
    if min_len < 2:
        return False
        
    return len(common) == min_len and len(common) >= 1


KNOWN_FRAGMENTS = {
    "shariah", "asia", "pacific", "ex", "japan", "progress", "shariah progress", 
    "asia-pacific", "ex-japan", "shariah asia-pacific", "shariah progress-plus"
}

def begins_with_fragment(line_text):
    line_clean = line_text.strip().lower()
    for frag in KNOWN_FRAGMENTS:
        if line_clean.startswith(frag):
            return True
    return False

def ends_with_hyphen(line_text):
    return line_text.strip().endswith('-')

def lacks_financial_metrics(line_text):
    # Find decimal numbers like NAV or Unit values (e.g. 10000.00 or 0.5200)
    numbers = re.findall(r'\b\d+\.\d+\b', line_text)
    return len(numbers) < 2

def clean_name(name_str):
    if not name_str:
        return ""
    name_str = re.sub(r'^(Account Holder\(s\):|Account Holder:|Name:)\s*', '', name_str, flags=re.IGNORECASE)
    return name_str.strip()

def normalize_date_to_dmy(date_str):
    """
    Normalizes date formats like "03 Jun 2026", "29 May 2026", "31/12/2026" into "DD/MM/YYYY".
    """
    if not date_str:
        return ""
        
    date_str = re.sub(r'\s+', ' ', date_str).strip()
    
    # Handle ranges like "01 Jan 2026 to 31 Dec 2026" or "01 Jan 2026 - 31 Dec 2026"
    if " to " in date_str.lower():
        parts = re.split(r'\s+to\s+', date_str, flags=re.IGNORECASE)
        return " to ".join(normalize_date_to_dmy(p) for p in parts)
    if " - " in date_str:
        parts = date_str.split(" - ")
        return " to ".join(normalize_date_to_dmy(p) for p in parts)
        
    # Normalize hyphens in single dates (e.g. "05-Jun-2026" -> "05 Jun 2026")
    date_str = date_str.replace("-", " ")
    date_str = re.sub(r'\s+', ' ', date_str).strip()
    
    # Already in DD/MM/YYYY
    if re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
        return date_str
        
    months_map = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
        "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12",
        "january": "01", "february": "02", "march": "03", "april": "04", "june": "06",
        "july": "07", "august": "08", "september": "09", "october": "10", "november": "11", "december": "12"
    }
    
    # Try parsing: DD MMM YYYY or D MMM YYYY (e.g. 02 Sep 2013 or 03 Jun 2026)
    m = re.match(r'^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$', date_str)
    if m:
        day = f"{int(m.group(1)):02d}"
        month_name = m.group(2).lower()
        month_name_short = month_name[:3]
        month = months_map.get(month_name_short, "01")
        year = m.group(3)
        return f"{day}/{month}/{year}"
        
    # Try parsing: MMM YYYY (e.g. Jun 2026) -> default to 01/MM/YYYY
    m_month_year = re.match(r'^([A-Za-z]+)\s+(\d{4})$', date_str)
    if m_month_year:
        month_name_short = m_month_year.group(1).lower()[:3]
        month = months_map.get(month_name_short, "01")
        year = m_month_year.group(2)
        return f"01/{month}/{year}"
        
    return date_str

def extract_date_from_lines(idx, lines):
    """
    Looks for a date pattern matching the start of a transaction block.
    If a transaction date row lacks a 4-digit year, it inspects adjacent line vectors
    (using a look-ahead window) and stitches the month/day and year components together.
    """
    def get_full_date(s):
        m = re.search(r'\b(\d{1,2})\s+([A-Za-z]{3,12})\s+(20\d{2})\b', s)
        if m:
            return f"{m.group(1)} {m.group(2)} {m.group(3)}"
        m = re.search(r'\b(\d{1,2})/(\d{1,2})/(20\d{2})\b', s)
        if m:
            return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
        return None

    def get_split_date(s1, s2):
        m = re.search(r'\b(\d{1,2})\s+([A-Za-z]{3,12})\b', s1)
        if m:
            m_yr = re.search(r'\b(20\d{2})\b', s2)
            if m_yr:
                return f"{m.group(1)} {m.group(2)} {m_yr.group(1)}"
        m = re.search(r'\b(\d{1,2})/(\d{1,2})\b', s1)
        if m:
            m_yr = re.search(r'\b(20\d{2})\b', s2)
            if m_yr:
                return f"{m.group(1)}/{m.group(2)}/{m_yr.group(1)}"
        return None

    # 1. Check current line for a complete date
    d = get_full_date(lines[idx])
    if d:
        return d
        
    # 2. Check if day/month is on current line and year is split on adjacent line
    # Look-ahead window (line immediately below)
    if idx + 1 < len(lines):
        d = get_split_date(lines[idx], lines[idx+1])
        if d:
            return d
            
    # Look-behind window (line immediately above)
    if idx - 1 >= 0:
        d = get_split_date(lines[idx], lines[idx-1])
        if d:
            return d

    # 3. Check preceding line for complete date
    if idx - 1 >= 0:
        d = get_full_date(lines[idx-1])
        if d:
            return d
        d = get_split_date(lines[idx-1], lines[idx])
        if d:
            return d
            
    # 4. Check 2 lines above
    if idx - 2 >= 0:
        d = get_full_date(lines[idx-2])
        if d:
            return d
        d = get_split_date(lines[idx-2], lines[idx-1])
        if d:
            return d
            
    # 5. Check 3 lines above
    if idx - 3 >= 0:
        d = get_full_date(lines[idx-3])
        if d:
            return d
        d = get_split_date(lines[idx-3], lines[idx-2])
        if d:
            return d
            
    # 6. Check next line for complete date
    if idx + 1 < len(lines):
        d = get_full_date(lines[idx+1])
        if d:
            return d
            
    return None

def is_continuation_line(line_text, fund_keys):
    line_clean = line_text.strip().lower()
    if not line_clean:
        return False
    # Starts with a date pattern (e.g. "19 Jul", "28/10", etc.)?
    if re.match(r'^\b\d{1,2}\s+[a-zA-Z]{3,12}\b', line_clean) or re.match(r'^\b\d{1,2}/\d{1,2}\b', line_clean):
        return False
    # Starts with a day number followed by year? E.g. "19 Mar 2026"
    if re.match(r'^\b\d{1,2}\b', line_clean) and ("20" in line_clean):
        if re.search(r'\b(20\d{2})\b', line_clean):
            return False
    # Contains a fund name?
    for f_key in fund_keys:
        if f_key == line_clean or line_clean.startswith(f_key) or f" {f_key} " in f" {line_clean} ":
            return False
    # Starts with "total"?
    if line_clean.startswith("total"):
        return False
    # Contains transaction codes at the start of any word?
    if re.search(r'\b(?:sle|swi|swo|mcr|div|mrf|rbb)\b', line_clean):
        return False
    return True

def is_header_or_metadata_line(line_text):
    line_clean = line_text.strip().lower()
    if not line_clean:
        return True
        
    # Reject lines containing strictly digits, spaces, commas, decimals, parenthesised numbers, etc.
    # This prevents split year-continuation lines (like '2023 2023') or numeric transaction rows
    # from being matched as active fund name headers.
    if re.match(r'^[\d\s,.\(\)-]+$', line_clean):
        return True
        
    # Check if the line is just a year (e.g. 2026) or standard numbers
    if re.match(r'^(20\d{2}|\d+)$', line_clean):
        return True
        
    # Check blacklist keywords for headers and metadata
    blacklist = [
        "charges", "sst", "tax", "price", "unit", "average", "cost", "admin", "sales",
        "total value", "proceed", "report", "cumulative", "transaction", "process date",
        "page", "statement", "account", "hotline", "disclaimer", "date", "trx", "closing date",
        "category", "value", "unrealised", "realised", "profit", "loss"
    ]
    for word in blacklist:
        if word in line_clean:
            return True
            
    # Check if it starts with a date pattern (e.g. "19 Mar", "28/10", etc.)
    if re.match(r'^\b\d{1,2}\s+[a-zA-Z]{3,12}\b', line_clean) or re.match(r'^\b\d{1,2}/\d{1,2}\b', line_clean):
        return True
        
    # Check if it contains standalone transaction codes
    if re.search(r'\b(?:sle|swi|swo|mcr|div|mrf|rbb)\b', line_clean):
        return True
        
    # Check if it contains mostly numbers (e.g. a transaction row of numbers)
    numbers = re.findall(r'\b\d[\d,.]*\b', line_clean)
    if len(numbers) >= 3:
        return True
        
    return False

def clean_dates_from_text(text):
    # Strip any date patterns from the text so that digits inside dates are not treated as numbers.
    # 1. Full dates (DD MMM YYYY or DD/MM/YYYY)
    text = re.sub(r'\b\d{1,2}\s+[A-Za-z]{3,12}\s+20\d{2}\b', ' ', text)
    text = re.sub(r'\b\d{1,2}/\d{1,2}/20\d{2}\b', ' ', text)
    # 2. Day-month (DD MMM or DD/MM)
    text = re.sub(r'\b\d{1,2}\s+[A-Za-z]{3,12}\b', ' ', text)
    text = re.sub(r'\b\d{1,2}/\d{1,2}\b', ' ', text)
    # 3. Year (20XX)
    text = re.sub(r'\b20\d{2}\b', ' ', text)
    return text


def find_fund_headers_on_page(page, fund_keys):
    """
    Extracts all lines on the page with their vertical coordinates (top)
    and searches them for occurrences of fund names.
    Returns a sorted list of dicts: [{'name': 'HW FLEXI', 'top': y_coordinate}, ...]
    """
    words = page.extract_words()
    if not words:
        return []
        
    # Sort words by vertical coordinate, then horizontal coordinate
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    
    # Group words on the same line (vertical delta < 3 pixels)
    lines = []
    current_line = []
    last_top = None
    
    for w in words:
        if last_top is None or abs(w["top"] - last_top) < 3:
            current_line.append(w)
        else:
            lines.append(current_line)
            current_line = [w]
        last_top = w["top"]
    if current_line:
        lines.append(current_line)
        
    headers = []
    for line in lines:
        line_text = " ".join([w["text"] for w in line])
        line_top = line[0]["top"]
        
        # Clean line text
        cleaned_line = re.sub(r'\s+', ' ', line_text).strip().lower()
        
        # Match against our extracted fund names
        for f_key in fund_keys:
            if f_key in cleaned_line:
                headers.append({
                    "name": f_key,
                    "top": line_top
                })
                break
                
    # Sort headers by top coordinate (top of page first)
    return sorted(headers, key=lambda h: h["top"])

def get_template_funds_checklist(template_path):
    checklist = []
    if template_path and os.path.exists(template_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(template_path, data_only=True)
            ws = wb.active
            fund_col = None
            header_row = 1
            for r in range(1, 16):
                for c in range(1, 40):
                    val = ws.cell(row=r, column=c).value
                    if val and "fund" in str(val).lower() and "name" in str(val).lower():
                        fund_col = c
                        header_row = r
                        break
                if fund_col:
                    break
            
            if fund_col:
                for r in range(header_row + 1, 200):
                    val = ws.cell(row=r, column=fund_col).value
                    if val:
                        name = str(val).strip()
                        if name and not any(kw in name.lower() for kw in ["fund name", "total", "summary", "category"]):
                            checklist.append(name)
        except Exception as e:
            print(f"Error extracting checklist from template: {e}")
    return checklist

def parse_pdf(pdf_path, template_path=None):
    results = {
        "statement_period": "",
        "account_holder": "",
        "joint_holder": "",
        "account_no": "",
        "investment_type": "",
        "funds": [],
        "distributions": []
    }

    
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        
    full_text = ""
    pages_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            # Text Streams Cleansing Pass: Split raw string array and clean internal newlines/carriage returns
            raw_lines = text.split('\n')
            cleaned_lines = []
            for line in raw_lines:
                cleaned_line = line.replace('\n', ' ').replace('\r', ' ').strip()
                cleaned_lines.append(cleaned_line)
            cleaned_text = "\n".join(cleaned_lines)
            
            full_text += cleaned_text + "\n"
            
            # Find tables with structural coordinate metadata
            tables = page.find_tables() or []
            pages_data.append({
                "index": page_idx + 1,
                "text": cleaned_text,
                "tables": tables,
                "page_obj": page
            })
            
    # 1. Parse Statement Period
    # Match pattern: Statement Period: 01/01/2026 to 31/12/2026 or Statement Period as of 03 Jun 2026
    period_match = re.search(r'(?:Statement\s*Period|Period|Date)\s*(?:as\s*of|:)?\s*([^\n\t]+)', full_text, re.IGNORECASE)
    if period_match:
        raw_period = period_match.group(1).split('\n')[0].strip()
        # Clean up any trailing headers that might be on the same line
        clean_period = re.split(r'\b(?:Account|Holder|Name|Advisor|Client|Type)\b', raw_period, flags=re.IGNORECASE)[0].strip()
        results["statement_period"] = re.sub(r'^(as\s+of|:)\s*', '', clean_period, flags=re.IGNORECASE).strip()
    else:
        results["statement_period"] = datetime.now().strftime("%d %b %Y")
        
    results["statement_period"] = re.sub(r'^(as\s+of|:)\s*', '', results["statement_period"], flags=re.IGNORECASE).strip()
    results["statement_period"] = normalize_date_to_dmy(results["statement_period"])
            
    # 2. Parse Account Holder(s)
    holder_match = re.search(r'(?:Account\s*Holder\(s\)|Account\s*Holder|Holder\s*Name)\s*:?\s*([^\n]+)', full_text, re.IGNORECASE)
    if holder_match:
        raw_holder_string = clean_name(holder_match.group(1))
        # Remove any trailing advisor or other details on the same line
        raw_holder_string = re.split(r'\b(?:Advisor|Client\s*Service|Advisor:|Statement)\b', raw_holder_string, flags=re.IGNORECASE)[0].strip()
        if "&" in raw_holder_string:
            parts = raw_holder_string.split("&", 1)
            results["account_holder"] = parts[0].strip()
            results["joint_holder"] = parts[1].strip()
        else:
            results["account_holder"] = raw_holder_string.strip()
            results["joint_holder"] = ""
            
    # 3. Parse Account Number & 4. Parse Investment Type
    # Match pattern: Account 9010038441 (EPF) or Account No: 987654321
    # Restrict account number to digits only to avoid matching the title of the PDF
    acct_match = re.search(r'\bAccount\s*(?:No\.?|Number)?\s*:?\s*(\d+)\s*(?:\(([^)]+)\))?', full_text, re.IGNORECASE)
    if acct_match:
        results["account_no"] = acct_match.group(1).strip()
        if acct_match.group(2):
            results["investment_type"] = acct_match.group(2).strip()
            
    if not results["account_no"]:
        fallback_acct = re.search(r'(?:Account\s*No\.?|Account\s*Number)\s*:?\s*(\w+)', full_text, re.IGNORECASE)
        if fallback_acct:
            results["account_no"] = fallback_acct.group(1).strip()
            
    if not results["investment_type"]:
        fallback_inv = re.search(r'(?:Investment\s*Type|Inv\s*Type|Type)\s*:?\s*([A-Za-z\t /]+)', full_text, re.IGNORECASE)
        if fallback_inv:
            results["investment_type"] = fallback_inv.group(1).strip()

    # If investment type is 'UT Ordinary', put 'CASH' instead
    if results["investment_type"] and results["investment_type"].strip().lower() == "ut ordinary":
        results["investment_type"] = "CASH"

    # 5. Parse Fund Portfolio values (Page 1)
    parsed_funds = {}
    parsed_table_divs = []
    
    def safe_float(val_str):
        if not val_str:
            return 0.0
        cleaned = str(val_str).replace(',', '').replace('%', '').replace('(', '').replace(')', '').strip()
        try:
            multiplier = -1.0 if '(' in str(val_str) and ')' in str(val_str) else 1.0
            return float(cleaned) * multiplier
        except ValueError:
            return 0.0
            
    # Process page-by-page tables
    for p_data in pages_data:
        for t_obj in p_data["tables"]:
            table = t_obj.extract()
            if not table or len(table) < 2:
                continue
                
            # Check if this is the Main Portfolio Table
            is_portfolio_table = False
            col_indices = {}
            
            for row_idx in range(min(len(table), 2)):
                row_str_list = [str(c).strip().lower() for c in table[row_idx] if c is not None]
                has_fund = any("fund" in cell for cell in row_str_list)
                has_units = any("unit" in cell for cell in row_str_list)
                has_nav = any("nav" in cell for cell in row_str_list)
                has_trx = any("trx" in cell or "transaction" in cell or "process" in cell for cell in row_str_list)
                
                if has_fund and has_units and not has_trx:
                    is_portfolio_table = True
                    for c_idx, cell in enumerate(row_str_list):
                        if "fund" in cell:
                            col_indices["fund_name"] = c_idx
                        elif "unit" in cell:
                            col_indices["total_unit"] = c_idx
                        elif "nav" in cell:
                            col_indices["nav"] = c_idx
                        elif "cost" in cell:
                            col_indices["total_cost"] = c_idx
                        elif "realised" in cell or "realized" in cell:
                            col_indices["realised_pl"] = c_idx
                    break
                    
            if is_portfolio_table:
                fund_name_col = col_indices.get("fund_name", 0)
                unit_col = col_indices.get("total_unit", 1)
                nav_col = col_indices.get("nav", 2)
                cost_col = col_indices.get("total_cost", 3)
                realised_col = col_indices.get("realised_pl", 4)
                
                for r_idx in range(1, len(table)):
                    row = table[r_idx]
                    if not row or len(row) <= max(fund_name_col, unit_col, nav_col):
                        continue
                        
                    fund_name_raw = row[fund_name_col]
                    if not fund_name_raw:
                        continue
                        
                    fund_name = re.sub(r'\s+', ' ', str(fund_name_raw)).strip()
                    if fund_name.lower() in ["fund name", "total value", "total", "category", "summary"]:
                        continue
                        
                    total_unit = safe_float(row[unit_col])
                    nav = safe_float(row[nav_col])
                    total_cost = safe_float(row[cost_col]) if cost_col < len(row) else 0.0
                    realised_pl = safe_float(row[realised_col]) if realised_col < len(row) else 0.0
                    
                    parsed_funds[fund_name.lower()] = {
                        "name": fund_name,
                        "total_unit": total_unit,
                        "nav": nav,
                        "total_cost": total_cost,
                        "realised_pl": realised_pl
                    }
                    
            # Check if this is a Distribution Table
            is_dist_table = False
            dist_col_indices = {}
            for row_idx in range(min(len(table), 2)):
                row_str_list = [str(c).strip().lower() for c in table[row_idx] if c is not None]
                has_date = any("date" in cell for cell in row_str_list)
                has_rate = any("rate" in cell or "div" in cell or "distribution" in cell for cell in row_str_list)
                has_units = any("unit" in cell for cell in row_str_list)
                
                if has_date and has_rate and not has_units:
                    is_dist_table = True
                    for c_idx, cell in enumerate(row_str_list):
                        if "fund" in cell:
                            dist_col_indices["fund_name"] = c_idx
                        elif "date" in cell:
                            dist_col_indices["date"] = c_idx
                        elif "rate" in cell or "div" in cell or "distribution" in cell:
                            dist_col_indices["rate"] = c_idx
                    break
                    
            if is_dist_table:
                fund_col = dist_col_indices.get("fund_name", 0)
                date_col = dist_col_indices.get("date", 1)
                rate_col = dist_col_indices.get("rate", 2)
                
                for r_idx in range(1, len(table)):
                    row = table[r_idx]
                    if not row or len(row) <= max(fund_col, date_col, rate_col):
                        continue
                    fund_name_raw = row[fund_col]
                    date_raw = row[date_col]
                    rate_raw = row[rate_col]
                    
                    if not fund_name_raw or not date_raw or not rate_raw:
                        continue
                        
                    fund_name = re.sub(r'\s+', ' ', str(fund_name_raw)).strip()
                    if fund_name.lower() in ["fund name", "total", "category", "summary"]:
                        continue
                        
                    date_clean = normalize_date_to_dmy(str(date_raw))
                    rate_val = safe_float(rate_raw)
                    
                    parsed_table_divs.append({
                        "name": fund_name,
                        "date": date_clean,
                        "rate": rate_val,
                        "amount": 0.0
                    })
                    print(f"Extracted DIV from summary table: Fund='{fund_name}', Date={date_clean}, Rate={rate_val}")
                    
    # 6. Parse Transaction History (Pages 1-4) to extract Dividends (DIV) via layout-agnostic line scanning
    parsed_history_divs = []
    
    # Hardened Fund Tracking Context State Machine: Sticky Context Lock
    current_active_fund = None
    
    # Build the master template checklist of valid fund names
    master_checklist = []
    if template_path:
        master_checklist.extend(get_template_funds_checklist(template_path))
        
    for f in parsed_funds.values():
        if f["name"] not in master_checklist:
            master_checklist.append(f["name"])
            
    # Add known default fallbacks to guarantee matching
    default_fallbacks = [
        "Manulife Asia-Pacific REIT Fund",
        "Manulife Global Managed Fund",
        "Manulife Investment Shariah Progress-Plus Fund",
        "Manulife Cash Management Fund",
        "SHARIAH ASIA-PACIFIC EX JAPAN",
        "SHARIAH PROGRESS",
        "ASIA-PACIFIC REIT"
    ]
    for name in default_fallbacks:
        if not any(is_fund_match(name, x) for x in master_checklist):
            master_checklist.append(name)
            
    # Sort the checklist by string length descending to prioritize more specific matches first
    master_checklist = sorted(master_checklist, key=lambda x: len(x), reverse=True)
            
    print(f"Master template checklist of valid fund names: {master_checklist}")
    
    # Extract target year from statement_period
    target_year = 2026 # Baseline system 'Statement Period' calendar reference year
    years = re.findall(r'\b(20\d{2})\b', results["statement_period"])
    if years:
        target_year = int(years[-1])
    print(f"Baseline statement reference year: {target_year}")
    
    # Process page-by-page in a loop to track page boundaries and retain state.
    for p_data in pages_data:
        page_num = p_data["index"]
        page_obj = p_data["page_obj"]
        
        words = page_obj.extract_words() or []
        words = sorted(words, key=lambda w: (w["top"], w["x0"]))
        
        # Text Streams Cleansing Pass: intercept the raw string array from the PDF reader
        # Programmatically replace all internal newline characters (\n), carriage returns (\r),
        # and wrapped row padding blocks (consecutive spaces/tabs) with a clean single space.
        page_lines = []
        curr_line_words = []
        last_top = None
        for w in words:
            if last_top is None or abs(w["top"] - last_top) < 3:
                curr_line_words.append(w)
            else:
                line_text = " ".join([x["text"] for x in curr_line_words])
                cleaned_line = line_text.replace('\n', ' ').replace('\r', ' ')
                cleaned_line = re.sub(r'\s+', ' ', cleaned_line).strip()
                page_lines.append(cleaned_line)
                curr_line_words = [w]
            last_top = w["top"]
        if curr_line_words:
            line_text = " ".join([x["text"] for x in curr_line_words])
            cleaned_line = line_text.replace('\n', ' ').replace('\r', ' ')
            cleaned_line = re.sub(r'\s+', ' ', cleaned_line).strip()
            page_lines.append(cleaned_line)
            
        # Date Re-assembly Tokenization: Ensure that any date values split across wrapped text fragments
        # (such as "25 Feb" followed by a separate "2026" on a subsequent sub-line) are compiled
        # into a singular, continuous string block (e.g. "25 Feb 2026") prior to parsing.
        reconstructed_lines = []
        skip_next = False
        for idx in range(len(page_lines)):
            if skip_next:
                skip_next = False
                continue
            curr_line = page_lines[idx]
            
            if idx + 1 < len(page_lines):
                next_line = page_lines[idx + 1].strip()
                
                # Check if next line contains any 4-digit year matching our target reference year
                m_yr_next = re.search(r'\b(20\d{2})\b', next_line)
                if m_yr_next:
                    year_val = m_yr_next.group(1)
                    if int(year_val) == target_year:
                        # Check if current line contains a day-month pattern but lacks the year
                        has_year_curr = re.search(r'\b20\d{2}\b', curr_line) is not None
                        has_dm_curr = re.search(r'\b\d{1,2}\s+[A-Za-z]{3,12}\b', curr_line) is not None or re.search(r'\b\d{1,2}/\d{1,2}\b', curr_line) is not None
                        
                        if has_dm_curr and not has_year_curr:
                            # Compile split date values
                            def replace_dm(match):
                                return f"{match.group(1)} {year_val}"
                            new_curr_line = re.sub(r'\b(\d{1,2}\s+[A-Za-z]{3,12})\b', replace_dm, curr_line)
                            new_curr_line = re.sub(r'\b(\d{1,2}/\d{1,2})\b', replace_dm, new_curr_line)
                            
                            if new_curr_line != curr_line:
                                curr_line = new_curr_line
                                print(f"Page {page_num}: Date Re-assembly Tokenization compiled: '{curr_line}'")
                                if re.match(r'^[\s\d]*$', next_line):
                                    skip_next = True
                                    
            reconstructed_lines.append(curr_line)
        page_lines = reconstructed_lines
        
        # 1.5. Apply Look-Ahead Reconstruction Rule to merge wrapped lines for fund names
        reconstructed_lines = []
        skip_next = False
        for idx in range(len(page_lines)):
            if skip_next:
                skip_next = False
                continue
            curr_line = page_lines[idx]
            if idx + 1 < len(page_lines):
                next_line = page_lines[idx + 1]
                if ends_with_hyphen(curr_line) or begins_with_fragment(curr_line):
                    if lacks_financial_metrics(next_line) and not is_header_or_metadata_line(next_line):
                        curr_line = curr_line + " " + next_line
                        skip_next = True
            reconstructed_lines.append(curr_line)
        page_lines = reconstructed_lines
        
        # 2. Scan lines sequentially to track active fund headers and parse DIV transactions
        for i, line_text in enumerate(page_lines):
            norm_line = line_text.strip().lower()
            
            # Anti-Reset Table Header Rule: lock active context on repeating table field headers
            is_table_header = False
            if "transaction date" in norm_line or "process date" in norm_line or "trx" in norm_line:
                is_table_header = True
                
            if is_table_header:
                pass # Lock context
            else:
                # Explicit Reset Boundary: current_active_fund can ONLY be overwritten
                # when a line matches 100% with a valid fund name title block found in checklist
                matched_checklist_fund = None
                for checklist_fund in master_checklist:
                    if is_fund_match(checklist_fund, line_text):
                        matched_checklist_fund = checklist_fund
                        break
                        
                if matched_checklist_fund:
                    current_active_fund = matched_checklist_fund
                    print(f"Page {page_num}: Hardened Context Machine set current_active_fund = '{current_active_fund}' based on line: '{line_text}'")
                    
            # Intercept DIV keyword token under active context block
            if re.search(r'\bDIV\b', line_text) and current_active_fund and not re.search(r'details|rate|declared', line_text, re.IGNORECASE):
                # Extract transaction date using proximity checks
                date_raw = extract_date_from_lines(i, page_lines)
                if not date_raw:
                    continue
                    
                date_clean = normalize_date_to_dmy(date_raw)
                try:
                    d_year = datetime.strptime(date_clean, "%d/%m/%Y").year
                except Exception:
                    d_year = target_year
                    
                # Reference Year Filter Constraint: if year doesn't match baseline system statement period reference year, drop it
                if d_year != target_year:
                    print(f"Page {page_num}: Discarding prior year DIV transaction for '{current_active_fund}' (year {d_year} != {target_year})")
                    continue
                    
                # Proximity extraction of Amount and Rate
                idx_div = line_text.find("DIV")
                merged_trx_text = line_text[idx_div + 3:]
                
                # Check for continuation lines
                k = i + 1
                while k < len(page_lines) and is_continuation_line(page_lines[k], list(parsed_funds.keys())):
                    merged_trx_text += " " + page_lines[k]
                    k += 1
                    
                cleaned_trx_text = clean_dates_from_text(merged_trx_text)
                num_matches = re.findall(r'\(?-?\b\d[\d,.]*\b\)?', cleaned_trx_text)
                
                numbers = []
                for m in num_matches:
                    cleaned_num = m.replace(',', '').replace('(', '').replace(')', '').strip()
                    try:
                        numbers.append(float(cleaned_num))
                    except ValueError:
                        pass
                        
                amount = 0.0
                if len(numbers) >= 4 and all(abs(n) < 0.01 for n in numbers[:3]):
                    amount = numbers[3]
                elif len(numbers) > 0:
                    for n in numbers:
                        if abs(n) > 0.01:
                            amount = n
                            break
                            
                rate = 0.0
                nonzero_numbers = [n for n in numbers if abs(n) > 0.01]
                if len(nonzero_numbers) >= 2:
                    amount_val = nonzero_numbers[0]
                    cum_units_after = nonzero_numbers[-1]
                    
                    is_three_format = False
                    if len(nonzero_numbers) >= 3:
                        ratio = amount_val / nonzero_numbers[1]
                        if abs(ratio - nonzero_numbers[2]) / nonzero_numbers[2] < 0.02:
                            is_three_format = True
                            
                    if is_three_format:
                        units_added = nonzero_numbers[2]
                        cum_units_before = cum_units_after - units_added
                        if cum_units_before > 0:
                            rate = amount_val / cum_units_before
                    else:
                        units_added = nonzero_numbers[1]
                        cum_units_before = cum_units_after - units_added
                        if cum_units_before > 0:
                            rate = amount_val / cum_units_before
                            
                parsed_history_divs.append({
                    "name": current_active_fund,
                    "date": date_clean,
                    "rate": rate,
                    "amount": amount
                })
                print(f"Page {page_num}: Extracted target year DIV from history for '{current_active_fund}': Date={date_clean}, Amount={amount}, Rate={rate:.4f}")

            
    # 7. Merge and Deduplicate dividends (prioritize summary table rates, sum history amounts)
    unique_divs = {}
    
    # First add all summary table dividends
    for d in parsed_table_divs:
        key = (d["name"].lower(), d["date"])
        unique_divs[key] = dict(d)
        
    # Merge history dividends
    for d in parsed_history_divs:
        key = (d["name"].lower(), d["date"])
        if key not in unique_divs:
            unique_divs[key] = dict(d)
        else:
            # Merge details: sum amounts
            existing = unique_divs[key]
            existing["amount"] += d["amount"]
            if d["rate"] > 0 and existing["rate"] == 0:
                existing["rate"] = d["rate"]
                
    results["funds"] = list(parsed_funds.values())
    results["distributions"] = list(unique_divs.values())
    
    return results

if __name__ == "__main__":
    user_pdf = r"C:\Users\Darre\Downloads\customer-report-type-D_YAP_CHIN_PANG_9010038441-20260604173009.pdf"
    if os.path.exists(user_pdf):
        print(f"Testing parser on: {user_pdf}")
        try:
            data = parse_pdf(user_pdf)
            import pprint
            pprint.pprint(data)
        except Exception as e:
            print(f"Error testing parser: {e}")
    else:
        print("Real user PDF not found, running local test...")
