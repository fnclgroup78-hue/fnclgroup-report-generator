import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def create_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Report"
    
    # Enable grid lines visibility
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1B365D")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="333333")
    formula_font = Font(name="Calibri", size=10, italic=True, color="000000")
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    data_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Title Block
    ws.merge_cells("A2:Y2")
    ws["A2"] = "MANULIFE CUSTOMER INVESTMENT SUMMARY"
    ws["A2"].font = title_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Headers rearranged per customer override:
    # Col A: Date
    # Col B: AVERAGE COST
    # Col C: Joint Holder
    # Col D: Account No
    # Col E: Investment Type
    # Col F: TOTAL AMOUNT INVESTED
    # Col G: TOTAL UNIT ACCUMULATED
    # Col H: Account Holder
    # Col I: NAV AS AT
    # Col J: Fund Name
    # Col K: CURRENT VALUE
    # Col L: PROFIT & LOSS
    # Col M: PERCENTAGE
    headers = [
        "Date", "AVERAGE COST", "Joint Holder", "Account No", "Investment Type",
        "TOTAL AMOUNT INVESTED", "TOTAL UNIT ACCUMULATED", "Account Holder", "NAV AS AT",
        "Fund Name", "CURRENT VALUE", "PROFIT & LOSS", "PERCENTAGE",
        "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
    ]
    
    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Funds to write
    funds = [
        "Manulife Asia-Pacific REIT Fund",
        "Manulife Global Managed Fund",
        "Manulife Investment Shariah Progress-Plus Fund",
        "Manulife Cash Management Fund"
    ]
    
    # Write template rows with fund names and formulas
    for row_idx, fund in enumerate(funds, 6):
        # Format the whole row cells
        for col_idx in range(1, 26):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
            # Shading alternate rows
            if row_idx % 2 == 0:
                cell.fill = data_fill
        
        # Write Fund Name to Column J (10)
        ws.cell(row=row_idx, column=10).value = fund
        ws.cell(row=row_idx, column=10).alignment = Alignment(horizontal="left")
        
        # Write pre-configured Excel formulas for calculations to test zero overrides
        # AVERAGE COST (Col B / 2) = TOTAL AMOUNT INVESTED (Col F / 6) / TOTAL UNIT ACCUMULATED (Col G / 7)
        cell_avg = ws.cell(row=row_idx, column=2)
        cell_avg.value = f"=F{row_idx}/G{row_idx}"
        cell_avg.font = formula_font
        cell_avg.number_format = "#,##0.0000"
        
        # CURRENT VALUE (Col K / 11) = TOTAL UNIT ACCUMULATED (Col G / 7) * NAV AS AT (Col I / 9)
        cell_curr = ws.cell(row=row_idx, column=11)
        cell_curr.value = f"=G{row_idx}*I{row_idx}"
        cell_curr.font = formula_font
        cell_curr.number_format = "#,##0.00"
        
        # PROFIT & LOSS (Col L / 12) = CURRENT VALUE (Col K / 11) - TOTAL AMOUNT INVESTED (Col F / 6)
        cell_pl = ws.cell(row=row_idx, column=12)
        cell_pl.value = f"=K{row_idx}-F{row_idx}"
        cell_pl.font = formula_font
        cell_pl.number_format = "#,##0.00"
        
        # PERCENTAGE (Col M / 13) = PROFIT & LOSS (Col L / 12) / TOTAL AMOUNT INVESTED (Col F / 6)
        cell_pct = ws.cell(row=row_idx, column=13)
        cell_pct.value = f"=L{row_idx}/F{row_idx}"
        cell_pct.font = formula_font
        cell_pct.number_format = "0.00%"
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    ws.row_dimensions[5].height = 28
    
    os.makedirs(os.path.dirname("backend/tests/template.xlsx"), exist_ok=True)
    wb.save("backend/tests/template.xlsx")
    print("Excel template created successfully at backend/tests/template.xlsx")

def create_pdf_statement():
    pdf_path = "backend/tests/manulife_statement.pdf"
    doc = SimpleDocTemplate(pdf_path, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1B365D'),
        spaceAfter=15
    )
    
    label_style = ParagraphStyle(
        'Label',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.HexColor('#333333')
    )
    
    val_style = ParagraphStyle(
        'Value',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#555555')
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#333333')
    )
    
    story = []
    
    # Title
    story.append(Paragraph("MANULIFE STATEMENT OF ACCOUNT", title_style))
    story.append(Spacer(1, 10))
    
    # Client & Account Info
    info_data = [
        [Paragraph("Statement Period:", label_style), Paragraph("01/01/2026 to 31/12/2026", val_style),
         Paragraph("Account Holder(s):", label_style), Paragraph("Darren Tan & Melissa Lim", val_style)],
        [Paragraph("Account No:", label_style), Paragraph("987654321", val_style),
         Paragraph("Investment Type:", label_style), Paragraph("Cash", val_style)]
    ]
    
    info_table = Table(info_data, colWidths=[110, 150, 110, 170])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Tabular Fund Metrics Title
    story.append(Paragraph("<b>Investment Portfolio Heuristics</b>", label_style))
    story.append(Spacer(1, 8))
    
    # Tabular Fund Metrics Table
    fund_headers = ["Fund Name", "Total Unit", "NAV (MYR)", "Total Cost", "Realised Profit / Loss"]
    fund_rows = [
        ["Manulife Asia-Pacific REIT Fund", "10000.00", "0.5200", "5000.00", "200.00"],
        ["Manulife Global Managed Fund", "5000.00", "1.1000", "5200.00", "-100.00"],
        ["Manulife Investment Shariah Progress-Plus Fund", "0.00", "0.4500", "3800.00", "0.00"]
    ]
    
    table_data = [[Paragraph(f"<b>{h}</b>", table_header_style if idx == 0 else table_header_style) for idx, h in enumerate(fund_headers)]]
    for row in fund_rows:
        table_data.append([
            Paragraph(row[0], table_cell_style),
            Paragraph(row[1], table_cell_style),
            Paragraph(row[2], table_cell_style),
            Paragraph(row[3], table_cell_style),
            Paragraph(row[4], table_cell_style)
        ])
        
    fund_table = Table(table_data, colWidths=[200, 80, 80, 90, 90])
    fund_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    
    story.append(fund_table)
    story.append(Spacer(1, 20))
    
    # Distributions Title
    story.append(Paragraph("<b>Distribution / DIV Details</b>", label_style))
    story.append(Spacer(1, 8))
    
    # Distributions table
    div_headers = ["Fund Name", "Distribution Date", "DIV Rate (MYR)"]
    div_rows = [
        ["Manulife Asia-Pacific REIT Fund", "31/03/2026", "0.0150"],
        ["Manulife Asia-Pacific REIT Fund", "30/09/2026", "0.0200"],
        ["Manulife Global Managed Fund", "15/06/2026", "0.0500"],
        ["Manulife Investment Shariah Progress-Plus Fund", "31/12/2026", "0.0300"]
    ]
    
    div_table_data = [[Paragraph(f"<b>{h}</b>", table_header_style) for h in div_headers]]
    for row in div_rows:
        div_table_data.append([
            Paragraph(row[0], table_cell_style),
            Paragraph(row[1], table_cell_style),
            Paragraph(row[2], table_cell_style)
        ])
        
    div_table = Table(div_table_data, colWidths=[240, 150, 150])
    div_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    
    story.append(div_table)
    story.append(Spacer(1, 20))
    
    # Transaction History Title
    story.append(Paragraph("<b>Transaction History</b>", label_style))
    story.append(Spacer(1, 8))
    
    # Fund header text line (must be outside the table!)
    story.append(Paragraph("<b>Manulife Asia-Pacific REIT Fund</b>", label_style))
    story.append(Spacer(1, 6))
    
    # Transaction Table columns
    trx_headers = ["Transaction Date", "Process Date", "Trx", "Amount", "Units", "Cumulative Units"]
    trx_rows = [
        ["15 Jun 2026", "17 Jun 2026", "SLE", "1000.00", "2000.00", "2000.00"],
        ["25 Feb", "26 Feb", "DIV", "1015.00", "2000.00", "4000.00"],
        ["2026", "", "", "", "", ""],
        ["24 Aug 2026", "25 Aug 2026", "DIV", "150.00", "300.00", "2300.00"],
        ["24 Aug 2026", "25 Aug 2026", "DIV", "200.00", "400.00", "2700.00"],
        ["31 Dec 2026", "02 Jan 2027", "DIV", "100.00", "200.00", "2900.00"],
        ["28 Oct 2025", "29 Oct 2025", "DIV", "50.00", "100.00", "1000.00"]
    ]
    
    trx_table_data = [[Paragraph(f"<b>{h}</b>", table_header_style) for h in trx_headers]]
    for row in trx_rows:
        trx_table_data.append([
            Paragraph(row[0], table_cell_style),
            Paragraph(row[1], table_cell_style),
            Paragraph(row[2], table_cell_style),
            Paragraph(row[3], table_cell_style),
            Paragraph(row[4], table_cell_style),
            Paragraph(row[5], table_cell_style)
        ])
        
    trx_table = Table(trx_table_data, colWidths=[95, 95, 60, 80, 80, 95])
    trx_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    story.append(trx_table)
    
    # Generate pages 2 to 18 as placeholders to test multi-page index limits
    for page_num in range(2, 19):
        story.append(PageBreak())
        story.append(Paragraph(f"<b>Page {page_num} - Placeholder</b>", label_style))
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"This is a placeholder for statement page {page_num}.", val_style))
        
    # Page 19 - Start of Shariah Progress transaction history (establishes active context)
    story.append(PageBreak())
    story.append(Paragraph("<b>Transaction History (Continued)</b>", label_style))
    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>Manulife Investment Shariah Progress-Plus Fund</b>", label_style))
    story.append(Spacer(1, 6))
    
    trx_headers_sp = ["Transaction Date", "Process Date", "Trx", "Amount", "Units", "Cumulative Units"]
    trx_rows_sp_p19 = [
        ["15 Jan 2026", "16 Jan 2026", "SLE", "1000.00", "2000.00", "2000.00"]
    ]
    
    table_data_sp_p19 = [[Paragraph(f"<b>{h}</b>", table_header_style) for h in trx_headers_sp]]
    for row in trx_rows_sp_p19:
        table_data_sp_p19.append([
            Paragraph(row[0], table_cell_style),
            Paragraph(row[1], table_cell_style),
            Paragraph(row[2], table_cell_style),
            Paragraph(row[3], table_cell_style),
            Paragraph(row[4], table_cell_style),
            Paragraph(row[5], table_cell_style)
        ])
    trx_table_sp_p19 = Table(table_data_sp_p19, colWidths=[95, 95, 60, 80, 80, 95])
    trx_table_sp_p19.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    story.append(trx_table_sp_p19)
    
    # Page 20 - Overflow of Shariah Progress transaction history without fund header
    story.append(PageBreak())
    story.append(Paragraph("<b>Transaction History (Continued)</b>", label_style))
    story.append(Spacer(1, 8))
    
    trx_rows_sp_p20 = [
        ["28 Apr 2026", "29 Apr 2026", "DIV", "1470.00", "3000.00", "5000.00"]
    ]
    
    table_data_sp_p20 = [[Paragraph(f"<b>{h}</b>", table_header_style) for h in trx_headers_sp]]
    for row in trx_rows_sp_p20:
        table_data_sp_p20.append([
            Paragraph(row[0], table_cell_style),
            Paragraph(row[1], table_cell_style),
            Paragraph(row[2], table_cell_style),
            Paragraph(row[3], table_cell_style),
            Paragraph(row[4], table_cell_style),
            Paragraph(row[5], table_cell_style)
        ])
    trx_table_sp_p20 = Table(table_data_sp_p20, colWidths=[95, 95, 60, 80, 80, 95])
    trx_table_sp_p20.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]))
    story.append(trx_table_sp_p20)
    
    doc.build(story)
    print("PDF statement created successfully at backend/tests/manulife_statement.pdf")

def create_single_column_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Report"
    ws.views.sheetView[0].showGridLines = True
    
    title_font = Font(name="Calibri", size=16, bold=True, color="1B365D")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="333333")
    formula_font = Font(name="Calibri", size=10, italic=True, color="000000")
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    data_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Title
    ws.merge_cells("A2:R2")
    ws["A2"] = "MANULIFE CUSTOMER INVESTMENT SUMMARY (SINGLE COL)"
    ws["A2"].font = title_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Headers
    headers_dict = {
        1: "Date",
        2: "AVERAGE COST",
        3: "Joint Holder",
        4: "Account No",
        5: "Investment Type",
        6: "TOTAL AMOUNT INVESTED",
        7: "TOTAL UNIT ACCUMULATED",
        8: "Account Holder",
        9: "NAV AS AT",
        10: "Fund Name",
        11: "CURRENT VALUE",
        12: "",
        13: "PROFIT & LOSS",
        14: "PERCENTAGE",
        15: "",
        16: "DISTRIBUTION DECLARED THIS YEAR",
        17: "DIS MONTH"
    }
    
    for c_idx, header in headers_dict.items():
        cell = ws.cell(row=5, column=c_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    funds = [
        "Manulife Asia-Pacific REIT Fund",
        "Manulife Global Managed Fund",
        "Manulife Investment Shariah Progress-Plus Fund",
        "Manulife Cash Management Fund"
    ]
    
    row_idx = 8
    
    # 1. Fund 1: Manulife Asia-Pacific REIT Fund (Row 8-12)
    months_to_write = ["Aug", "Sep", "Dec", "Jan", "Apr", "Mar"]
    for m in months_to_write:
        for c in range(1, 18):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = data_fill
                
        # Write Fund Name only on the first row (Row 8)
        if m == "Aug":
            ws.cell(row=row_idx, column=10).value = "Manulife Asia-Pacific REIT Fund"
            
        ws.cell(row=row_idx, column=17).value = m # DIS MONTH
        
        # Formulas
        ws.cell(row=row_idx, column=2).value = f"=F{row_idx}/G{row_idx}" # AVERAGE COST
        ws.cell(row=row_idx, column=11).value = f"=G{row_idx}*I{row_idx}" # CURRENT VALUE
        ws.cell(row=row_idx, column=13).value = f"=K{row_idx}-F{row_idx}" # PROFIT & LOSS
        ws.cell(row=row_idx, column=14).value = f"=M{row_idx}/F{row_idx}" # PERCENTAGE
        
        ws.cell(row=row_idx, column=2).font = formula_font
        ws.cell(row=row_idx, column=11).font = formula_font
        ws.cell(row=row_idx, column=13).font = formula_font
        ws.cell(row=row_idx, column=14).font = formula_font
        
        ws.cell(row=row_idx, column=2).number_format = "#,##0.0000"
        ws.cell(row=row_idx, column=11).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=13).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=14).number_format = "0.00%"
        
        row_idx += 1
        
    # 2. Fund 2: Manulife Global Managed Fund (Row 13)
    for c in range(1, 18):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = data_font
        cell.border = thin_border
        
    ws.cell(row=row_idx, column=10).value = "Manulife Global Managed Fund"
    ws.cell(row=row_idx, column=17).value = "Jun"
    ws.cell(row=row_idx, column=2).value = f"=F{row_idx}/G{row_idx}"
    ws.cell(row=row_idx, column=11).value = f"=G{row_idx}*I{row_idx}"
    ws.cell(row=row_idx, column=13).value = f"=K{row_idx}-F{row_idx}"
    ws.cell(row=row_idx, column=14).value = f"=M{row_idx}/F{row_idx}"
    
    ws.cell(row=row_idx, column=2).number_format = "#,##0.0000"
    ws.cell(row=row_idx, column=11).number_format = "#,##0.00"
    ws.cell(row=row_idx, column=13).number_format = "#,##0.00"
    ws.cell(row=row_idx, column=14).number_format = "0.00%"
    
    row_idx += 1
    
    # 3. Fund 3: Manulife Investment Shariah Progress-Plus Fund (Row 14)
    for c in range(1, 18):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = data_font
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = data_fill
            
    ws.cell(row=row_idx, column=10).value = "Manulife Investment Shariah Progress-Plus Fund"
    ws.cell(row=row_idx, column=17).value = "May"
    ws.cell(row=row_idx, column=2).value = f"=F{row_idx}/G{row_idx}"
    ws.cell(row=row_idx, column=11).value = f"=G{row_idx}*I{row_idx}"
    ws.cell(row=row_idx, column=13).value = f"=K{row_idx}-F{row_idx}"
    ws.cell(row=row_idx, column=14).value = f"=M{row_idx}/F{row_idx}"
    
    ws.cell(row=row_idx, column=2).number_format = "#,##0.0000"
    ws.cell(row=row_idx, column=11).number_format = "#,##0.00"
    ws.cell(row=row_idx, column=13).number_format = "#,##0.00"
    ws.cell(row=row_idx, column=14).number_format = "0.00%"
    
    row_idx += 1
    
    # Auto adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    wb.save("backend/tests/template_single_col.xlsx")
    print("Excel single-column template created successfully at backend/tests/template_single_col.xlsx")

if __name__ == "__main__":
    create_excel_template()
    create_single_column_template()
    create_pdf_statement()
