import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import openpyxl
from parser import parse_pdf
from excel_writer import update_excel_report

pdf_path = r"C:\Users\Darre\OneDrive\Desktop\customer-report-type-D_NG_AIK_SIEW_5205411052-20260606154108.pdf"
template_path = r"C:\Users\Darre\Downloads\Generated_Report.xlsx"
if not os.path.exists(template_path):
    template_path = r"C:\Users\Darre\Downloads\Generated_Report (4).xlsx"
backup_path = template_path.replace(".xlsx", "_Backup.xlsx")

def inspect_and_process():
    print(f"Checking if input files exist...")
    print(f"PDF exists: {os.path.exists(pdf_path)}")
    print(f"Excel template exists: {os.path.exists(template_path)}")
    
    if not os.path.exists(pdf_path) or not os.path.exists(template_path):
        print("Error: Input files missing!")
        return

    # Backup the original template before modification
    if not os.path.exists(backup_path):
        import shutil
        shutil.copyfile(template_path, backup_path)
        print(f"Created backup of Excel template at: {backup_path}")

    # Inspect template sheet layout
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    print(f"\n--- Excel Template Inspection ---")
    print(f"Active Sheet Name: {ws.title}")
    print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
    
    # Let's print out the first 10 rows and 20 columns
    print("\nFirst 10 rows header and client info:")
    for r in range(1, 16):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 35)]
        if any(v is not None for v in row_vals):
            # Print row index and non-empty values
            non_empty = {c: v for c, v in enumerate(row_vals, 1) if v is not None}
            print(f"Row {r:02d}: {non_empty}")
            
    # Step 1: Parse PDF
    print(f"\n--- Step 1: Parsing PDF '{os.path.basename(pdf_path)}' ---")
    parsed_data = parse_pdf(pdf_path, template_path)
    
    print("\n--- Parsed Portfolio Summary Funds ---")
    for fund in parsed_data.get("funds", []):
        print(fund)
        
    print("\n--- Parsed Dividend/Distribution History ---")
    distributions = parsed_data.get("distributions", [])
    print(f"Total distributions parsed: {len(distributions)}")
    for d in sorted(distributions, key=lambda x: (x['name'], x['date'])):
        print(f"Fund: {d['name']:<40} Date: {d['date']:<12} Rate: {d['rate']:<10} Amount: {d['amount']}")
        
    # Check specific targets in parsed data
    # Shariah Progress (May rollover: date >= 25 Apr or month 5)
    # Al-Fauzan (April rollover)
    # Asia-Pacific REIT (March rollover)
    print("\n--- Targeted Check on Parsed Distributions ---")
    for name in ["SHARIAH PROGRESS", "AL-FAUZAN", "ASIA-PACIFIC REIT"]:
        matches = [d for d in distributions if name.lower() in d['name'].lower()]
        print(f"\nMatches for '{name}':")
        for m in matches:
            print(m)

    # Step 2: Run update_excel_report
    print(f"\n--- Step 2: Running update_excel_report ---")
    output_path = template_path.replace(".xlsx", "_Output.xlsx")
    matched_count = update_excel_report(template_path, output_path, parsed_data)
    print(f"Completed update. Matched count: {matched_count} (Saved to {output_path})")
    
    # Step 3: Programmatic Verification
    print(f"\n--- Step 3: Verifying Results in updated Excel ---")
    wb_updated = openpyxl.load_workbook(output_path, data_only=True)
    ws_updated = wb_updated.active
    
    print("\nVerifying specific target cells:")
    # Let's inspect rows 26, 33, 19 and print columns 1-25
    for r in [26, 33, 19]:
        row_vals = {c: ws_updated.cell(row=r, column=c).value for c in range(1, 26)}
        print(f"Row {r}: {row_vals}")
        
    # Let's run assertions
    header_row = 5 # default, but get_header_row_and_mapping finds it
    from excel_writer import get_header_row_and_mapping
    hr, col_map = get_header_row_and_mapping(ws_updated)
    is_single = ("dist_month" in col_map or "dist_rate" in col_map)
    print(f"Detected Layout: {'Single-Column' if is_single else 'Multi-Column'}")
    print(f"Column Mapping: {col_map}")
    
    if is_single:
        # Single column layout: DIV amount is in column P (16) and Month in column Q (17)
        # Let's verify ASIA-PACIFIC REIT at row 26
        val_26_month = ws_updated.cell(row=26, column=17).value
        val_26_amount = ws_updated.cell(row=26, column=16).value
        print(f"Row 26 (ASIA-PACIFIC REIT) Month: {val_26_month}, Amount: {val_26_amount}")
        assert str(val_26_month).strip().lower() == "mar", f"Row 26 month is not 'Mar', got '{val_26_month}'"
        assert abs(float(val_26_amount or 0) - 1015.00) < 0.01, f"Row 26 amount is not 1015.00, got '{val_26_amount}'"
        
        # Row 33 (SHARIAH PROGRESS) -> May (Month 5) contains 1470.00
        val_33_month = ws_updated.cell(row=33, column=17).value
        val_33_amount = ws_updated.cell(row=33, column=16).value
        print(f"Row 33 (SHARIAH PROGRESS) Month: {val_33_month}, Amount: {val_33_amount}")
        assert str(val_33_month).strip().lower() == "may", f"Row 33 month is not 'May', got '{val_33_month}'"
        assert abs(float(val_33_amount or 0) - 1470.00) < 0.01, f"Row 33 amount is not 1470.00, got '{val_33_amount}'"
        
        # Row 19 (AL-FAUZAN) -> April (Month 4) contains 855.00
        val_19_month = ws_updated.cell(row=19, column=17).value
        val_19_amount = ws_updated.cell(row=19, column=16).value
        print(f"Row 19 (AL-FAUZAN) Month: {val_19_month}, Amount: {val_19_amount}")
        assert str(val_19_month).strip().lower() == "apr", f"Row 19 month is not 'Apr', got '{val_19_month}'"
        assert abs(float(val_19_amount or 0) - 855.00) < 0.01, f"Row 19 amount is not 855.00, got '{val_19_amount}'"
    else:
        # Multi-column layout: Month columns are mapped in col_map
        mar_col = col_map.get("mar")
        may_col = col_map.get("may")
        apr_col = col_map.get("apr")
        
        # Row 24 (ASIA-PACIFIC REIT) -> March
        val_24 = ws_updated.cell(row=24, column=mar_col).value
        print(f"Row 24 (ASIA-PACIFIC REIT) March (col {mar_col}): {val_24}")
        assert abs(float(val_24 or 0) - 1015.00) < 0.01, f"Row 24 March is not 1015.00, got '{val_24}'"
        
        # Row 31 (SHARIAH PROGRESS) -> May
        val_31 = ws_updated.cell(row=31, column=may_col).value
        print(f"Row 31 (SHARIAH PROGRESS) May (col {may_col}): {val_31}")
        assert abs(float(val_31 or 0) - 1470.00) < 0.01, f"Row 31 May is not 1470.00, got '{val_31}'"
        
        # Row 33 (AL-FAUZAN) -> April
        val_33 = ws_updated.cell(row=33, column=apr_col).value
        print(f"Row 33 (AL-FAUZAN) April (col {apr_col}): {val_33}")
        assert abs(float(val_33 or 0) - 855.00) < 0.01, f"Row 33 April is not 855.00, got '{val_33}'"
        
    print("\n=======================================================")
    print(" VERIFICATION SUCCESSFUL: ALL CHECKS PASS FOR NG AIK SIEW!")
    print("=======================================================")

if __name__ == "__main__":
    inspect_and_process()
