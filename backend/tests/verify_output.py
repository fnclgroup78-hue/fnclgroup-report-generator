import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import openpyxl

def verify_output_excel(filepath):
    print(f"Starting programmatic verification of output file: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=False) # Load with formulas intact
    ws = wb.active
    
    # 1. Verify Client details mapped correctly across matched rows (rows 6, 7, 8)
    for r in [6, 7, 8]:
        assert ws.cell(row=r, column=1).value == "01/01/2026 to 31/12/2026", f"Row {r}: Statement Period mismatch"
        assert ws.cell(row=r, column=3).value == "Melissa Lim", f"Row {r}: Joint Holder mismatch"
        assert ws.cell(row=r, column=4).value == "987654321", f"Row {r}: Account No mismatch"
        assert ws.cell(row=r, column=5).value == "Cash", f"Row {r}: Investment Type mismatch"
        assert ws.cell(row=r, column=8).value == "Darren Tan", f"Row {r}: Account Holder mismatch"
        
    print("[PASS] Client header details correctly mapped to all rows.")

    # 2. Verify Fund 1: Manulife Asia-Pacific REIT Fund (Row 6)
    # TOTAL UNIT ACCUMULATED (Column G / 7) = 10000.00
    assert ws.cell(row=6, column=7).value == 10000.00, "Row 6: Total Unit mismatch"
    # NAV AS AT (Column I / 9) = 0.5200
    assert ws.cell(row=6, column=9).value == 0.5200, "Row 6: NAV mismatch"
    # TOTAL AMOUNT INVESTED (Column F / 6) = Cost (5000.00) - Realised P/L (200.00) = 4800.00
    assert ws.cell(row=6, column=6).value == 4800.00, "Row 6: Total Invested calculation mismatch"
    
    # 3. Verify Formula Preservation (B6, K6, L6, M6 formulas must be preserved)
    assert ws.cell(row=6, column=2).value == "=F6/G6", "Row 6: Average Cost formula overwritten"
    assert ws.cell(row=6, column=11).value == "=G6*I6", "Row 6: Current Value formula overwritten"
    assert ws.cell(row=6, column=12).value == "=K6-F6", "Row 6: Profit/Loss formula overwritten"
    assert ws.cell(row=6, column=13).value == "=L6/F6", "Row 6: Percentage formula overwritten"
    print("[PASS] Pre-existing formulas protected and preserved.")
    
    # 4. Verify Month Roll-Over Rule for Dividends (Fund 1, Row 6)
    # March 31 -> Rolls over to April (Month 4 / Column Q)
    assert ws.cell(row=6, column=17).value == 0.0150, "Row 6: March 31 DIV (0.015) failed to rollover to April (Column Q)"
    # Sept 30 -> Rolls over to October (Month 10 / Column W)
    assert ws.cell(row=6, column=23).value == 0.0200, "Row 6: Sept 30 DIV (0.020) failed to rollover to October (Column W)"
    # 25 Feb -> Rolls over to March (Month 3 / Column P)
    assert ws.cell(row=6, column=16).value == 1015.00, "Row 6: 25 Feb DIV (1015.00) failed to rollover to March (Column P)"
    
    # 5. Verify Month Roll-Over Rule for Dividends (Fund 2, Row 7)
    # June 15 -> No rollover, stays in June (Month 6 / Column S)
    assert ws.cell(row=7, column=19).value == 0.0500, "Row 7: June 15 DIV (0.050) failed to map to June (Column S)"
    
    # 6. Verify Month Roll-Over Rule for Dividends (Fund 3, Row 8)
    # Dec 31 -> Rolls over to January of next year (Month 1 / Column N)
    assert ws.cell(row=8, column=14).value == 0.0300, "Row 8: Dec 31 DIV (0.030) failed to rollover to January (Column N)"
    # 28 Apr -> Rolls over to May (Month 5 / Column R)
    assert ws.cell(row=8, column=18).value == 1470.00, "Row 8: 28 Apr DIV (1470.00) failed to rollover to May (Column R)"
    print("[PASS] Month roll-over rule and dividend column mapping fully verified.")
    
    # 8. Verify Zero Unit Rule: Row 8 has units = 0.00, so NAV and Percentage must be overwritten and display as "-"
    assert ws.cell(row=8, column=7).value == 0.00, "Row 8: Total Unit should be 0.00"
    assert ws.cell(row=8, column=9).value == 0.00, "Row 8: NAV should be 0.00 when units are 0.00"
    # PERCENTAGE is Column M (13) in multi-column layout
    assert ws.cell(row=8, column=13).value == "=L8/F8", "Row 8: Percentage formula should be preserved"
    assert ws.cell(row=8, column=9).number_format == '#,##0.0000;-#,##0.0000;"-"', "Row 8: NAV number format should render zero as '-'"
    assert ws.cell(row=8, column=13).number_format == '0.00%;-0.00%;"-"', "Row 8: Percentage number format should render zero as '-'"
    print("[PASS] Zero-units NAV and Percentage override mapping verified successfully.")
    
    # 7. Style Preservation Check (fonts and borders)
    # Font name check on Fund Name (Column J / 10)
    cell_font = ws.cell(row=6, column=10).font
    assert cell_font.name == "Calibri", "Font styling corrupted"
    # Row 9 (Manulife Cash Management Fund) should remain completely unmodified
    assert ws.cell(row=9, column=1).value is None, "Row 9 (unmatched) was incorrectly written to"
    assert ws.cell(row=9, column=7).value is None, "Row 9 (unmatched) values incorrectly populated"
    print("[PASS] Style preservation & zero override cell guardrails fully verified.")
    
    print("\n=======================================================")
    print(" PROGRAMMATIC VERIFICATION SUCCESSFUL: 100% CORRECT!")
    print("=======================================================")

def verify_single_column():
    print("\nStarting programmatic verification of Single-Column output...")
    from excel_writer import update_excel_report
    from parser import parse_pdf
    
    # 1. Process files
    data = parse_pdf("backend/tests/manulife_statement.pdf", "backend/tests/template_single_col.xlsx")
    output_path = "backend/tests/output_single_col_result.xlsx"
    update_excel_report("backend/tests/template_single_col.xlsx", output_path, data)
    
    # 2. Load generated file
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb.active
    
    # 3. Verify mappings and rollover values
    # Match REIT fund (Rows 8-13)
    # Row 8: Month Aug -> Stays in August (150.00 + 200.00 same-date sum = 350.00)
    assert ws.cell(row=8, column=16).value == 350.00, f"Row 8: August DIV Amount mismatch (got {ws.cell(row=8, column=16).value})"
    # Row 9: Month Sep -> No transaction rolled over to September now, remains empty (None)
    assert ws.cell(row=9, column=16).value is None, f"Row 9: September DIV Amount mismatch (got {ws.cell(row=9, column=16).value})"
    # Row 10: Month Dec -> Stays at None (empty/blank)
    assert ws.cell(row=10, column=16).value is None, f"Row 10: December DIV Amount mismatch (got {ws.cell(row=10, column=16).value})"
    # Row 11: Month Jan -> Rolled over from Dec 31 DIV (100.00)
    assert ws.cell(row=11, column=16).value == 100.00, f"Row 11: January DIV Amount mismatch (got {ws.cell(row=11, column=16).value})"
    # Row 12: Month Apr -> Rolled over from March 31 DIV (summary rate 0.0150 -> amount defaults to 0.0 -> remains None)
    assert ws.cell(row=12, column=16).value is None, f"Row 12: April DIV Amount mismatch (got {ws.cell(row=12, column=16).value})"
    # Row 13: Month Mar -> Rolled over from 25 Feb DIV (1015.00)
    assert ws.cell(row=13, column=16).value == 1015.00, f"Row 13: March DIV Amount mismatch (got {ws.cell(row=13, column=16).value})"
    
    # Row 15: Shariah Progress Month May -> Rolled over from 28 Apr DIV (1470.00)
    assert ws.cell(row=15, column=16).value == 1470.00, f"Row 15: May DIV Amount mismatch (got {ws.cell(row=15, column=16).value})"
    
    # Check that formats are correct
    assert ws.cell(row=8, column=16).number_format == '#,##0.00;-#,##0.00;"-"'
    
    print("[PASS] Single-column DIV Amount mapping, 25th rollover, and target year filters fully verified.")
    print("\n=======================================================")
    print(" SINGLE COLUMN VERIFICATION SUCCESSFUL: 100% CORRECT!")
    print("=======================================================")

def verify_normalization_rules():
    print("\nStarting programmatic verification of Normalization & Matching Rules...")
    from parser import is_fund_match, normalize_fund_name
    
    # Test normalization
    assert normalize_fund_name("SHARIAH ASIA-\nPACIFIC EX JAPAN") == "SHARIAH ASIA PACIFIC EX JAPAN"
    assert normalize_fund_name("ASIA-PACIFIC EX JAPAN FUND") == "ASIA-PACIFIC EX JAPAN FUND"
    assert normalize_fund_name("SHARIAH\nPROGRESS") == "SHARIAH PROGRESS"
    assert normalize_fund_name("  -SHARIAH- -PROGRESS-  ") == "SHARIAH PROGRESS"
    
    # Test normalized token matching
    assert is_fund_match("SHARIAH ASIA-PACIFIC EX JAPAN", "SHARIAH ASIA-\nPACIFIC EX JAPAN")
    assert is_fund_match("SHARIAH ASIA-PACIFIC EX JAPAN", "shariah-asia-pacific-ex-japan")
    assert is_fund_match("ASIA-PACIFIC EX JAPAN FUND", "asia-pacific-ex-japan-fund")
    assert is_fund_match("SHARIAH PROGRESS", "SHARIAH\nPROGRESS")
    assert is_fund_match("SHARIAH PROGRESS", "shariah-progress")
    assert is_fund_match("SHARIAH PROGRESS", "  Shariah   Progress  ")
    
    print("[PASS] Normalization and token matching rules verified successfully.")
    print("\n=======================================================")
    print(" NORMALIZATION VERIFICATION SUCCESSFUL: 100% CORRECT!")
    print("=======================================================")

def verify_interpage_persistence():
    print("\nStarting programmatic verification of Inter-Page Persistence & Rollovers...")
    from parser import is_fund_match
    
    # 1. Check subset token match for Shariah Progress
    excel_name = "Manulife Investment Shariah Progress-Plus Fund"
    pdf_name = "SHARIAH PROGRESS"
    assert is_fund_match(excel_name, pdf_name), "SHARIAH PROGRESS subset matching failed!"
    
    # 2. Check 28 Apr month rollover shift to May (Month 5)
    from excel_writer import apply_month_rollover
    assert apply_month_rollover("28/04/2026") == 5, "28 Apr failed to roll over to May!"
    
    # 3. Verify that the 28 Apr 2026 DIV row is bound to Shariah Progress and amount 1470.00 is in May column
    wb = openpyxl.load_workbook("backend/tests/output_result.xlsx", data_only=False)
    ws = wb.active
    
    # Shariah Progress is on row 8, May is column 18 (R)
    val_may = ws.cell(row=8, column=18).value
    assert val_may == 1470.00, f"Row 8: May column (18) should be 1470.00, got {val_may}"
    
    print("[PASS] Inter-page persistence rules and matching verified.")
    print("\n=======================================================")
    print(" PERSISTENCE VERIFICATION SUCCESSFUL: 100% CORRECT!")
    print("=======================================================")

if __name__ == "__main__":
    from generate_samples import create_excel_template, create_single_column_template, create_pdf_statement
    from excel_writer import update_excel_report
    from parser import parse_pdf
    
    print("Regenerating test templates and PDF statement...")
    create_excel_template()
    create_single_column_template()
    create_pdf_statement()
    
    print("Generating multi-column output...")
    data = parse_pdf("backend/tests/manulife_statement.pdf", "backend/tests/template.xlsx")
    update_excel_report("backend/tests/template.xlsx", "backend/tests/output_result.xlsx", data)
    
    verify_output_excel("backend/tests/output_result.xlsx")
    verify_single_column()
    verify_normalization_rules()
    verify_interpage_persistence()
